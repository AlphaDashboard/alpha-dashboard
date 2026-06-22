import openpyxl
from datetime import datetime
from django.db import transaction
from dashboard.models import Voucher, VoucherFact, AccountMaster

class VoucherImportService:
    """
    Service class to handle the business logic of importing vouchers from an Excel file.
    """
    
    @staticmethod
    def parse_and_validate_excel(excel_file):
        """
        Parses the Excel file and extracts raw voxel data.
        Returns a dict of voucher data or raises ValueError.
        """
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise ValueError('The uploaded file is empty or missing headers.')

        headers = [str(h).strip().lower() for h in rows[0] if h is not None]
        expected_headers = ['voucher_no', 'voucher_date', 'remarks', 'type', 'account_master', 'amount']
        
        for req_header in expected_headers:
            if req_header not in headers:
                raise ValueError(f'Missing required column in Excel: {req_header}')
        
        header_map = {h: i for i, h in enumerate(headers)}
        vouchers_data = {}
        
        for idx, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
                
            voucher_no = str(row[header_map['voucher_no']]).strip()
            if not voucher_no or voucher_no == 'None':
                raise ValueError(f'Row {idx}: Missing voucher_no')
                
            if voucher_no not in vouchers_data:
                vouchers_data[voucher_no] = {
                    'date': row[header_map['voucher_date']],
                    'remarks': row[header_map['remarks']],
                    'facts': []
                }
            
            fact_type = str(row[header_map['type']]).strip().upper()
            alpha_code = str(row[header_map['account_master']]).strip()
            amount_val = row[header_map['amount']]
            
            try:
                amount = float(amount_val or 0)
            except ValueError:
                raise ValueError(f'Row {idx}: Invalid amount format')
                
            vouchers_data[voucher_no]['facts'].append({
                'row_idx': idx,
                'type': fact_type,
                'account_master': alpha_code,
                'amount': amount
            })
            
        return vouchers_data

    @staticmethod
    def process_imported_vouchers(vouchers_data):
        """
        Processes the grouped voucher dictionary and persists to the database.
        Runs inside an atomic transaction.
        """
        from dashboard.models.cashbank import CashBank
        from dashboard.services.transaction_sp_helper import execute_sp_manage_transaction
        
        with transaction.atomic():
            for v_no, v_data in vouchers_data.items():
                if CashBank.objects.filter(voucher_no=v_no).exists():
                    raise ValueError(f"Duplicate voucher number: '{v_no}' already exists.")
                    
                total_a = sum(f['amount'] for f in v_data['facts'] if f['type'] == 'A')
                total_b = sum(f['amount'] for f in v_data['facts'] if f['type'] == 'B')
                
                if abs(total_a - total_b) > 0.001:
                    raise ValueError(f"Voucher '{v_no}' Error: Voucher A and B must be equal (Total A: {total_a}, Total B: {total_b}).")
                    
                if total_a == 0 and total_b == 0:
                    raise ValueError(f"Voucher '{v_no}': Cannot import zero-value voucher.")
                
                v_date = v_data['date']
                if not v_date:
                    raise ValueError(f"Voucher '{v_no}': Missing voucher date.")
                
                if isinstance(v_date, str):
                    try:
                        v_date = datetime.strptime(v_date, '%Y-%m-%d').date()
                    except ValueError:
                        raise ValueError(f"Voucher '{v_no}': Invalid date format. Use YYYY-MM-DD.")
                elif hasattr(v_date, 'date'):
                    v_date = v_date.date()
                
                header_data = {
                    'voucher_no': v_no,
                    'date': v_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'tran_type': 'J000',
                    'rpid': None,
                    'amount': total_a,
                    'remarks': v_data['remarks'] if v_data['remarks'] and str(v_data['remarks']) != 'None' else ''
                }
                
                detail_rows = []
                for fact in v_data['facts']:
                    if fact['type'] not in ['A', 'B']:
                         raise ValueError(f"Row {fact['row_idx']}: Invalid Fact Type '{fact['type']}'. Must be A or B.")
                         
                    account_master = None
                    try:
                        val = int(fact['account_master'])
                        account_master = AccountMaster.objects.filter(groupID=val).first()
                    except ValueError:
                        pass
                    if not account_master:
                        account_master = AccountMaster.objects.filter(Account_Name__iexact=fact['account_master']).first()
                        
                    if not account_master:
                        raise ValueError(f"Row {fact['row_idx']} AccountMaster group not found: '{fact['account_master']}'.")
                        
                    detail_rows.append({
                        'account_master': account_master.pk,
                        'amount': fact['amount'],
                        'remarks': '',
                        'rpid': fact['type']
                    })
                
                execute_sp_manage_transaction('INSERT', 'SECTION_A', header_data, detail_rows, 'system')
        return len(vouchers_data)

class VoucherExportService:
    """
    Handles extracting Voucher objects into an Excel object.
    """
    @staticmethod
    def generate_excel(queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vouchers Export"
        
        headers = ["Voucher No", "Voucher Date", "Remarks", "Amount", "Status"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        for voucher in queryset:
            amount = float(voucher.total_amount or 0)
            status = 'Active' if voucher.is_active else 'Inactive'
            ws.append([
                str(voucher.voucher_number),
                voucher.voucher_date.strftime('%Y-%m-%d'),
                str(voucher.remarks or ''),
                amount,
                status
            ])
            
        return wb
